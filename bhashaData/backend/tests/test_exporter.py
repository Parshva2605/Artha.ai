from __future__ import annotations

from unittest.mock import MagicMock

import os

try:
	from backend.pipeline.exporter import (
		EXPORT_COLUMNS,
		generate_metadata,
		export_csv,
		export_excel,
		export_huggingface,
		export_json,
		export_parquet,
		prepare_rows_for_export,
		run_export_pipeline,
	)
except ModuleNotFoundError:
	from pipeline.exporter import (
		EXPORT_COLUMNS,
		generate_metadata,
		export_csv,
		export_excel,
		export_huggingface,
		export_json,
		export_parquet,
		prepare_rows_for_export,
		run_export_pipeline,
	)


def make_test_rows(n: int = 5) -> list[dict]:
	return [
		{
			"text_original": f"Test text {i} original",
			"text_clean": f"Test text {i} clean",
			"language": "hi",
			"language_name": "Hindi",
			"script": "devanagari",
			"domain": "social_media",
			"source": "reddit",
			"source_url": None,
			"source_subreddit": "india",
			"label_sentiment": "positive",
			"label_topic": None,
			"label_ner": None,
			"confidence": 0.90,
			"confidence_reason": "Positive tone",
			"llm_used": "ollama",
			"needs_review": False,
			"app_id": None,
			"star_rating": None,
			"rating_hint": None,
		}
		for i in range(1, n + 1)
	]


def test_prepare_rows_for_export_adds_id_and_job_id() -> None:
	rows = make_test_rows(3)
	prepared = prepare_rows_for_export(rows, "test-job-1")
	assert prepared[0]["id"] == 1
	assert prepared[2]["id"] == 3
	assert all(row["job_id"] == "test-job-1" for row in prepared)


def test_csv_export_creates_file_with_correct_columns() -> None:
	rows = prepare_rows_for_export(make_test_rows(), "j1")
	path = export_csv(rows, "/tmp/test_export.csv")
	import pandas as pd

	df = pd.read_csv(path)
	for column in EXPORT_COLUMNS:
		assert column in df.columns


def test_json_export_preserves_indian_characters() -> None:
	rows = prepare_rows_for_export(
		[
			{
				**make_test_rows(1)[0],
				"text_original": "यह बहुत अच्छा है",
			}
		],
		"j2",
	)
	path = export_json(rows, "/tmp/test_export.json")
	import json

	with open(path, encoding="utf-8") as file_handle:
		data = json.load(file_handle)
	assert "यह बहुत अच्छा है" in data[0]["text_original"]


def test_excel_export_creates_file_with_dataset_sheet() -> None:
	rows = prepare_rows_for_export(make_test_rows(), "j3")
	path = export_excel(rows, "/tmp/test_export.xlsx")
	import openpyxl

	workbook = openpyxl.load_workbook(path)
	assert "Dataset" in workbook.sheetnames
	assert "Quality_Info" in workbook.sheetnames


def test_parquet_export_creates_readable_file() -> None:
	rows = prepare_rows_for_export(make_test_rows(), "j4")
	path = export_parquet(rows, "/tmp/test_export.parquet")
	import pandas as pd

	df = pd.read_parquet(path)
	assert len(df) == 5
	assert str(df["confidence"].dtype) == "float64"


def test_huggingface_export_creates_folder_with_required_files() -> None:
	rows = prepare_rows_for_export(make_test_rows(), "j5")
	path = export_huggingface(rows, "/tmp/test_hf_export")
	assert os.path.isdir(path)
	files = os.listdir(path)
	assert any("arrow" in filename for filename in files)


def test_generate_metadata_includes_all_required_keys() -> None:
	rows = prepare_rows_for_export(make_test_rows(), "j6")
	mock_report = MagicMock()
	mock_report.label_distribution = {}
	mock_report.per_language_distribution = {}
	mock_report.balance_result.is_balanced = True
	mock_report.per_language_quality = {"hi": 90.0}
	mock_report.overall_quality_score = 90.0
	mock_report.benchmark_comparison.benchmark_note = "test"
	mock_report.shortfall_warnings = []
	mock_report.low_quality_warning = None
	mock_report.claude_count = 0
	mock_report.openai_count = 0
	mock_report.ollama_count = 5
	mock_report.needs_review_count = 0
	metadata = generate_metadata(
		job_id="j6",
		rows=rows,
		quality_report=mock_report,
		requested_per_language={"hi": 5},
		label_type="sentiment",
		export_formats=["csv", "json"],
		domain="social_media",
	)
	assert metadata["platform"] == "Artha AI v1.0"
	assert metadata["total_rows"] == 5
	assert "hi" in metadata["per_language"]
	assert "overall" in metadata["quality_scores"]


def test_run_export_pipeline_raises_value_error_for_empty_format_list() -> None:
	rows = prepare_rows_for_export(make_test_rows(), "j7")
	mock_report = MagicMock()
	try:
		run_export_pipeline(
			rows=rows,
			job_id="j7",
			export_formats=[],
			quality_report=mock_report,
			requested_per_language={"hi": 5},
			label_type="sentiment",
			domain="social_media",
		)
		assert False, "Should have raised ValueError"
	except ValueError:
		pass